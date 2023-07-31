import openpyxl

save_path = '엑셀자동화/auto.xlsx'

wb = openpyxl.load_workbook(save_path)

ws = wb.active

ws['A1'] = '수하인명'
ws['C1'] = '수하인주소'
ws['D1'] = '수하인전화번호'
ws['E1'] = '수하인핸드폰번호'
ws['F1'] = '택배수량'
ws['G1'] = '택배운임'
ws['H1'] = '운임구분'
ws['I1'] = '품목명'
ws['K1'] = '배송메세지'

ws.cell(row = 2, column = 1, value = '홍길동')
ws.cell(row = 2, column = 3, value = '서울 용산구 한강로3가 16-49')
ws.cell(row = 2, column = 4, value = '02-3415-1111')
ws.cell(row = 2, column = 5, value = '010-0000-0000')
ws.cell(row = 2, column = 6, value = '1')
ws.cell(row = 2, column = 7, value = '2500')
ws.cell(row = 2, column = 8, value = '010')
ws.cell(row = 2, column = 9, value = '테스트품명 첫번째')
ws.cell(row = 2, column = 11, value = '친절 배송 부탁드립니다.')

ws.append(['이순신', '', '서울 용산구 청파로 40 삼구빌딩', '02-3415-2222', '010-0000-0000', 
            '2', '6000', '010', '테스트품명 두번째', '', '친절 배송 부탁드립니다.'])

wb.save(save_path)