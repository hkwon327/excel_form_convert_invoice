import openpyxl

total_wb = openpyxl.Workbook()

total_ws = total_wb.active

total_ws.title = 'data'

file_list = ['11번가1', '스마트스토어', 'ESM']

for file in file_list:
    wb = openpyxl.load_workbook(f'엑셀자동화/2023-07-29/{file}.xlsx', data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row = 2):
        data = []
        for cell in row:
            data.append(cell.value)
        total_ws.append(data)

total_wb.save('엑셀자동화/2023-07-29/total.xlsx')