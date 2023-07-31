import openpyxl

save_path = '엑셀자동화/2023-07-28/11번가.xlsx'

wb = openpyxl.load_workbook(save_path)

ws = wb['Sheet1']


for row in ws.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end = " ")
    print()


for x in range(2, ws.max_row + 1):
    for y in range(1, 13 + 1):
        print(ws.cell(row=x, column=y).value, end = " ")
    print()

for row in ws.iter_rows():
    print(row)
    