# open an excel file
import openpyxl


wb = openpyxl.load_workbook('엑셀자동화/2023-07-29/11번가.xlsx')

# select a sheet
ws = wb['Sheet1']

# edit cells
ws.delete_cols(1) #주문번호
ws.delete_cols(2) #주문금액
ws.delete_cols(3) #배송방법
ws.delete_cols(10) #서산간배송비


ws.insert_cols(1) #수화주명
ws.insert_cols(1) #우편번호
ws.insert_cols(1) #주소
ws.insert_cols(1) #수하인전화
ws.insert_cols(1) #수하인핸드폰번호
ws.insert_cols(1) #택배수량
ws.insert_cols(1) #택배운임_빈행
ws.insert_cols(1) #선착불_빈행
ws.insert_cols(1) #물품명
ws.insert_cols(1) #선택안함_빈행
ws.insert_cols(1) #배송메시지


max_row = ws.max_row

ws.move_range(f'M1:M{max_row}', cols = -12) #수화주명
ws.move_range(f'Q1:R{max_row}', cols = -15) #우편번호, 주소
ws.move_range(f'P1:P{max_row}', cols = -12) #수하인전화번호
ws.move_range(f'O1:O{max_row}', cols = -10) #수하인휴대폰번호
ws.move_range(f'T1:T{max_row}', cols = -14) #택배수량
ws.move_range(f'L1:L{max_row}', cols = -5) #물품명
ws.move_range(f'S1:S{max_row}', cols = -11) #배송메시지

ws.delete_cols(14) #배송비
ws.insert_cols(7) #택배운임_빈행
ws.insert_cols(7) #선착불_빈행
ws.insert_cols(10) #선택안함_빈행


for i in range(2, max_row+1):
    ws.cell(row=i, column=19, value='공구모아')
    ws.cell(row=i, column=20, value='경기 안양시 동안구 호계동, 555-9 국제유통단지18동127호')
    ws.cell(row=i, column=21, value='031-479-0988')


# save 
wb.save('엑셀자동화/2023-07-29/11번가1.xlsx')